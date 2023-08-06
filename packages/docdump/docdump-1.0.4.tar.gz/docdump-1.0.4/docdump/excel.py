from openpyxl import load_workbook

class excel_reader:
	def __init__(self):
		pass

	def read(self, path):
		'''
		Opens a excel file, extracts text and data

		inputs: str, a path to a PDF file
		outputs: str, a string of all the text in the pdf that could be extracted.
		'''
		try:
			workbook = load_workbook(filename=path, data_only=True) 
			#data only ensure that the value of the formula is imported not the formula itself
		except ValueError:
			raise ValueError("Workbook could not be loaded: {0}".format(path))
		
		text_dump = []

		for sheet_index in range(len(workbook.sheetnames)):
			workbook.active = sheet_index
			sheet = workbook.active
			for col in sheet["A:ZZ"]:
				for cell in col:
					if cell.value:
						text_dump.append(str(cell.value))
		
		return " ".join(text_dump)