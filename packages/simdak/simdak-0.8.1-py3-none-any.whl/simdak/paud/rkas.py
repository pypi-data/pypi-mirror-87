from __future__ import annotations
from bs4 import BeautifulSoup, Tag
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from requests import Session
from typing import Union
from typing import Dict, List, Optional, Type, TypeVar, TYPE_CHECKING

from . import (
    BaseSimdakPaud,
    JENIS_KOMPONEN,
    JENIS_PENGGUNAAN,
    PENGGUNAAN,
    get_key,
    get_fuzz,
)


class RkasData(BaseSimdakPaud):
    # TODO : Refactor!
    INDEX = "A"
    MAPPING = {
        "jenis_komponen_id": "B",
        "jenis_penggunaan_id": "C",
        "jenisbelanja": "D",
        "qty": "E",
        "satuan": "F",
        "hargasatuan": "G",
        "data_id": "H",
    }

    def __init__(
        self,
        jenis_komponen_id: int,
        jenis_penggunaan_id: int,
        jenisbelanja: str,
        qty: int,
        satuan: str,
        hargasatuan: int,
        data_id: str = "",
    ):
        self.jenis_komponen_id = int(jenis_komponen_id)
        self.jenis_penggunaan_id = int(jenis_penggunaan_id)
        self.jenisbelanja = jenisbelanja
        self.qty = int(qty)
        self.satuan = satuan
        self.hargasatuan = int(hargasatuan)
        if data_id and "=" in data_id:
            data_id = data_id.split("=")[-1]
        self.data_id = data_id
        self._logger.debug(f"RPD [{self}]")

    def update(self, **kwargs) -> RkasData:
        for k, v in kwargs.items():
            if hasattr(self, k):
                setattr(self, k, v)
        data = self.as_data()
        params = {"r": "boppaudrkas/update", "id": self.data_id}
        res = self._session.post(self._base_url, data, params=params)
        if not res.ok:
            self._logger.warning(f"Gagal mengupdate data [{self.data_id}]")
        return self

    def delete(self) -> bool:
        params = {"r": "boppaudrkas/delete", "id": self.data_id}
        res = self._session.get(self._base_url, params=params)
        return res.ok

    def as_data(self, **kwargs) -> dict:
        if self.jenis_komponen_id not in JENIS_PENGGUNAAN:
            raise ValueError(f"jenis komponen {self.jenis_komponen_id} tidak valid!")
        if self.jenis_penggunaan_id not in JENIS_PENGGUNAAN[self.jenis_komponen_id]:
            raise ValueError(
                f"jenis penggunaan {self.jenis_penggunaan_id} tidak valid!"
            )
        data = {
            "Boppaudrkas[jenis_komponen_id]": self.jenis_komponen_id,
            "Boppaudrkas[jenis_penggunaan_id]": self.jenis_penggunaan_id,
            "Boppaudrkas[jenisbelanja]": self.jenisbelanja.replace(" ", "+"),
            "Boppaudrkas[qty]": self.qty,
            "Boppaudrkas[satuan]": self.satuan.replace(" ", "+"),
            "Boppaudrkas[hargasatuan]": self.hargasatuan,
        }
        data.update(kwargs)
        return data

    def as_dict(self, **kwargs) -> dict:
        return {
            "jenis_komponen_id": self.jenis_komponen_id,
            "jenis_penggunaan_id": self.jenis_penggunaan_id,
            "jenisbelanja": self.jenisbelanja,
            "qty": self.qty,
            "satuan": self.satuan,
            "hargasatuan": self.hargasatuan,
            "data_id": self.data_id,
        }

    def to_row(self, ws: Union[Worksheet, Workbook], row: int):
        data = self.as_dict()
        try:
            for k, v in data.items():
                col = self.MAPPING[k]
                ws[f"{col}{row}"] = v
            # TODO : Buat ini dynamic
            ws[f"B{row}"] = JENIS_KOMPONEN.get(self.jenis_komponen_id)
            ws[f"C{row}"] = PENGGUNAAN.get(self.jenis_penggunaan_id)
        except Exception as e:
            self._logger.exception(
                f"Gagal memasukkan data ke baris {row-1}, karena {e}"
            )

    def __eq__(self, other: object) -> bool:
        if not isinstance(other, RkasData):
            return super().__eq__(other)
        return (
            self.jenis_komponen_id == other.jenis_komponen_id
            and self.jenis_penggunaan_id == other.jenis_penggunaan_id
            and self.jenisbelanja == other.jenisbelanja
            and self.qty == other.qty
            and self.satuan == other.satuan
            and self.hargasatuan == other.hargasatuan
        )

    def __str__(self) -> str:
        s = [
            f"JK: {JENIS_KOMPONEN[self.jenis_komponen_id]}",
            f"JP: {PENGGUNAAN[self.jenis_penggunaan_id]}",
            f"JB: {self.jenisbelanja}",
            f"Jumlah: {self.qty} {self.satuan}",
            f"Harga: {self.hargasatuan}",
        ]
        return ", ".join(s)

    @classmethod
    def from_tr(cls, tr: Tag) -> RkasData:
        tds = tr.findAll("td")
        if not tds:
            raise ValueError("Data tidak ditemukan")
        jenis_komponen_id = get_key(tds[2].get_text(), JENIS_KOMPONEN)
        if not jenis_komponen_id:
            raise ValueError(f"Jenis Komponen {jenis_komponen_id} tidak valid")
        jenis_penggunaan_id = get_key(
            tds[3].get_text(), JENIS_PENGGUNAAN[jenis_komponen_id]
        )
        if not jenis_penggunaan_id:
            raise ValueError(f"Jenis Penggunaan {jenis_penggunaan_id} tidak valid")
        return cls(
            jenis_komponen_id=jenis_komponen_id,
            jenis_penggunaan_id=jenis_penggunaan_id,
            jenisbelanja=tds[4].get_text(),
            qty=int(tds[5].get_text()),
            satuan=tds[6].get_text(),
            hargasatuan=int(tds[7].get_text()),
            data_id=tds[10].find("a")["href"] if len(tds) == 11 else "",
        )

    @classmethod
    def from_row(cls, ws: Union[Worksheet, Workbook], row: int) -> RkasData:
        data = {}
        for k, v in cls.MAPPING.items():
            data[k] = ws[f"{v}{row}"].value
        col = cls.MAPPING["jenis_komponen_id"]
        val = ws[f"{col}{row}"].value
        data["jenis_komponen_id"] = get_fuzz(val, JENIS_KOMPONEN, 1)
        col = cls.MAPPING["jenis_penggunaan_id"]
        val = ws[f"{col}{row}"].value
        data["jenis_penggunaan_id"] = get_fuzz(val, PENGGUNAAN, 21)
        return cls(**data)


class Rkas(BaseSimdakPaud):
    def __init__(
        self,
        no: str,
        npsn: str,
        satuan_pendidikan: str,
        alamat: str,
        alokasi: str,
        kegiatan_pembelajaran_dan_bermain: str,
        kegiatan_pendukung: str,
        kegiatan_lainnya: str,
        jumlah: str,
        url: str,
    ):
        self.no = no
        self.npsn = npsn
        self.satuan_pendidikan = satuan_pendidikan
        self.alamat = alamat
        self.alokasi = alokasi
        self.kegiatan_pembelajaran_dan_bermain = kegiatan_pembelajaran_dan_bermain
        self.kegiatan_pendukung = kegiatan_pendukung
        self.kegiatan_lainnya = kegiatan_lainnya
        self.jumlah = jumlah
        self.url = url
        self.id = self.url.split("&")[1].split("=")[-1]
        self._logger.info(f"RKAS ({self.id}) berhasil dibuka")

    def __call__(
        self, semester_id: int = 20201, save_as: Type[RkasData] = RkasData
    ) -> List[RkasData]:
        return self.get(semester_id, save_as)

    def get(
        self, semester_id: int = 20201, save_as: Type[RkasData] = RkasData
    ) -> List[RkasData]:
        results: List[RkasData] = []
        params = {"r": "boppaudrkas/create", "id": self.id, "semester_id": semester_id}
        res = self._session.get(self._base_url, params=params)
        if not res.ok:
            return results
        soup = BeautifulSoup(res.text, "html.parser")
        table: Tag = soup.findAll("table")[1]
        for tr in table.findAll("tr"):
            try:
                result = save_as.from_tr(tr)
                results.append(result)
            except ValueError as e:
                self._logger.exception(e)
                continue
        self._logger.info(f"Berhasil mendapatkan {len(results)} rpd")
        return results

    def create(
        self, rkas_data: RkasData, semester_id: int = 20201
    ) -> Optional[RkasData]:
        data = rkas_data.as_data()
        data.update({"yt0": "Simpan"})
        params = {"r": "boppaudrkas/create", "id": self.id, "semester_id": semester_id}
        res = self._session.post(self._base_url, data=data, params=params)
        if not res.ok:
            return None
        soup = BeautifulSoup(res.text, "html.parser")
        table: List[Tag] = soup.findAll("table")
        tr: Tag = table[-1].findAll("tr")[-1]
        return RkasData.from_tr(tr)


class SimdakRkasPaud(BaseSimdakPaud):
    # TODO : Refactor!
    def __call__(self, semester_id: int = 20201) -> List[Rkas]:
        return self.get(semester_id)

    def get(self, semester_id: int = 20201) -> List[Rkas]:
        results: List[Rkas] = []
        params = {
            "r": "boppaudrkas/index",
            "Boppaudrkas[semester_id]": semester_id,
            "yt0": "Cari",
        }
        res = self._session.get(self._base_url, params=params)
        if not res.status_code == 200:
            return results
        soup = BeautifulSoup(res.text, "html.parser")
        for tr in soup.findAll("tr", {"class": "view"}):
            tds = tr.findAll("td")
            result = Rkas(
                no=tds[0].get_text(),
                npsn=tds[1].get_text(),
                satuan_pendidikan=tds[2].get_text(),
                alamat=tds[3].get_text(),
                alokasi=tds[4].get_text(),
                kegiatan_pembelajaran_dan_bermain=tds[5].get_text(),
                kegiatan_pendukung=tds[6].get_text(),
                kegiatan_lainnya=tds[7].get_text(),
                jumlah=tds[8].get_text(),
                url=self._domain + tds[9].find("a")["href"] or "",
            )
            results.append(result)
        return results
