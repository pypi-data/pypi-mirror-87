import cv2 as cv
import openpyxl
import img2pdf
from PIL import Image
import os
class Cert_gen:
    def check():
        return "Its Oke :)"
    def generate(format,template_path,details_path,output_path,prefix_name='cert_',style={'font_size':1.5,'font_color':(0,0,0),'x':15,'y':7}):
        """
        Required :
        - format = png or pdf (String)
        - template_path = certificate template path - image (String)
        - details_path = excell file contains your certificate list
        - output path = path to create and save certificate
        Additional :
        - prefix_name = file name prefix(String)
        - Styling = {'font_size':1.5,'font_color':(0,0,0),'x':15,'y':7}
        Simple Usage Example :
            Cert_gen.Cert_gen.generate('png','tem.jpg','Daftar.xlsx','sertif')
        Custom Usage Example :
            Cert_gen.Cert_gen.generate('png','tem.jpg','Daftar.xlsx','sertif','cert101',{'font_size':1.5,'font_color':(0,0,0),'x':15,'y':7})
        """
        font_size = style['font_size']
        font_color = style['font_color']
        coordinate_y_adjustment = style['y']
        coordinate_x_adjustment = style['x']

        obj = openpyxl.load_workbook(details_path)
        sheet = obj.active
        for i in range(1,sheet.max_row):
            get_name = sheet.cell(row = i ,column = 1)
            get_owner = sheet.cell(row = i ,column = 2)
            certi_name = get_name.value
            certi_owner = get_owner.value
            img = cv.imread(template_path)
            font = cv.FONT_HERSHEY_SIMPLEX
            text_size = cv.getTextSize(certi_name, font, font_size, 10)[0]
            text_x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_adjustment
            text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_adjustment
            text_x = int(text_x)
            text_y = int(text_y)
            cv.putText(img, certi_name,
                    (text_x ,text_y ),
                    font,
                    font_size,
                    font_color, 2)
            if format == "png":
                certi_path = output_path +'/'+ prefix_name +str(certi_owner)+ '.png'
                cv.imwrite(certi_path,img)
            else:
                certi_path = output_path +'/temp'+ prefix_name +str(certi_owner)+ '.png'
                cv.imwrite(certi_path,img)
                pdf_path = output_path +'/'+ prefix_name +str(certi_owner)+ '.pdf'
                image = Image.open(certi_path)
                pdf_bytes = img2pdf.convert(image.filename)
                file = open(pdf_path, "wb")
                file.write(pdf_bytes)
                image.close()
                file.close()
                os.remove(certi_path)
