from __future__ import annotations
import os
from logging import getLogger
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Optional, Union
from . import RkasData, SimdakPaud
from simdak.template import TEMPLATE_FILE

CWD = os.getcwd()
COL_INDEX = RkasData.INDEX
COL_ID = RkasData.MAPPING.get("data_id")


def find_one(datas: List[RkasData], rpd: str) -> Optional[RkasData]:
    if not rpd or not datas:
        return None
    for d in datas:
        if d.data_id == rpd:
            return d
    return None


def exports(
    email: str, password: str, filename: str = "", sheet: str = "Sheet1"
) -> None:
    logger = getLogger("paud-export")
    simdak = SimdakPaud(email, password)
    rkas = simdak.rkas()[0]
    logger.info(f"Berhasil masuk akun {email}, mendapatkan data RPD...")
    rkas_datas = rkas.get(save_as=RkasData)
    filename = filename or f"{rkas.npsn}-Simdak-Paud"
    logger.info(f"Mengeksport {len(rkas_datas)} data RPD ke [{filename}]...")
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    filepath = os.path.join(CWD, filename)
    ws: Worksheet = None
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active
    ws.title = sheet
    for i, rkas_data in enumerate(rkas_datas):
        rkas_data.to_row(ws, i + 2)
        ws[f"{COL_INDEX}{i+2}"] = i + 1
    wb.save(filepath)
    logger.info(f"Berhasil mengeksport data ke {filename}!")


def imports(
    email: str,
    password: str,
    filename: str,
    start: int = 1,
    ke: int = 0,
    sheet: str = "Sheet1",
    header: bool = True,
    save: bool = True,
) -> None:
    logger = getLogger("paud-import")
    simdak = SimdakPaud(email, password)
    rkas = simdak.rkas()[0]
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    filepath = os.path.join(CWD, filename)
    ws: Worksheet = None
    if os.path.isfile(filepath):
        wb = load_workbook(filepath)
        sheets = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheet) if sheet in sheets else wb.active
    rkas_datas = rkas.get(save_as=RkasData)
    row = start + 1 if header else start
    imported = 0
    while True:
        if row == ke:
            break
        elif not ws[f"{COL_INDEX}{row}"].value:
            break
        data = RkasData.from_row(ws, row)
        old = find_one(rkas_datas, data.data_id)
        result: Optional[RkasData] = None
        if old:
            if old == data:
                logger.debug(f"{row} dilewati")
                row += 1
                continue
            logger.debug(f"{row} diperbarui")
            result = old.update(**data.as_dict())
        else:
            logger.debug(f"{row} dibuat")
            result = rkas.create(data)
        if result:
            ws[f"{COL_ID}{row}"] = result.data_id
        row += 1
        imported += 1
    logger.info(f"Berhasil memasukan data sebanyak {imported}")
    if save:
        wb.save(filepath)
    logger.info(f"Berhasil menyimpan data terbaru")
