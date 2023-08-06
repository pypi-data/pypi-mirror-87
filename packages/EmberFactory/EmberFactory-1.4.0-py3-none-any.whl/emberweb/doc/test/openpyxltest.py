import openpyxl
book = openpyxl.load_workbook("/Users/marbaix/wk/BurningEmbers/EmberData/embers-collection/main/SPM-no-additional-data/SROCC_SPM3.xlsx")
print("Worksheet name(s): {0}".format(book.sheetnames))
sh = book.worksheets[0]
sh = book['Graph parameters']
for row in sh.rows:
    print(row[0].value, row[1].value)
print(sh.cell(1,1).value)
print(sh.cell(2,1).value)
print(sh.cell(3,1).value)
print(sh.cell(4,1).value)
print(sh.cell(5,1).value)
print(sh.cell(6,1).value)
print(sh.cell(7,1).value)

# https://stackoverflow.com/questions/64264563/attributeerror-elementtree-object-has-no-attribute-getiterator-when-trying/65131301#65131301


