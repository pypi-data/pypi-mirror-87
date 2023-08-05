import os

from openpyxl import load_workbook
from docxtpl import DocxTemplate

from easytrans.constant import FAILURE_STATUS, SUCCESS_STATUS, DOC_FORMAT, XLS_FORMAT
from easytrans.utils import FileUtil


class DataExportDoc(object):

    @staticmethod
    def data_to_doc(data, docx_template_path, doc_dir=None, doc_name=None, variable_name="data"):
        try:
            if not data or not docx_template_path:
                return FAILURE_STATUS, "'data' or 'docx_template_path' error!"

            if not os.path.isfile(docx_template_path) or not os.path.splitext(docx_template_path)[1] in DOC_FORMAT:
                return FAILURE_STATUS, "'{0}' file format error!".format(docx_template_path)

            doc_t = DocxTemplate(docx_template_path)
            doc_t.render({variable_name: data})

            if not doc_name:
                doc_name = FileUtil.generate_file_name()

            if not doc_dir:
                doc_dir = os.getcwd()

            if not os.path.exists(doc_dir):
                os.makedirs(doc_dir)

            doc_path = os.path.join(doc_dir, "{0}.docx".format(doc_name))

            doc_t.save(doc_path)

            return SUCCESS_STATUS, doc_path

        except Exception as e:
            return FAILURE_STATUS, repr(e)


class DataExportExcel(object):

    @staticmethod
    def data_to_xls(data_list, xlsx_template_path, start_row=1, start_col=1, xls_dir=None, xls_name=None,
                    sheet_index=0):
        try:
            if not data_list or not xlsx_template_path:
                return FAILURE_STATUS, "'data_list' or 'xlsx_template_path' error!"

            if not os.path.isfile(xlsx_template_path) or not os.path.splitext(xlsx_template_path)[1] in XLS_FORMAT:
                return FAILURE_STATUS, "'{0}' file format error!".format(xlsx_template_path)

            wb = load_workbook(filename=xlsx_template_path)
            names = wb.sheetnames
            ws = wb.get_sheet_by_name(names[sheet_index])

            if not xls_name:
                xls_name = FileUtil.generate_file_name()

            if not xls_dir:
                xls_dir = os.getcwd()

            if not os.path.exists(xls_dir):
                os.makedirs(xls_dir)

            for row in range(len(data_list)):
                for col in range(len(data_list[row])):
                    ws.cell(row=start_row + row, column=start_col + col).value = data_list[row][col]

            xls_path = os.path.join(xls_dir, "{0}.xlsx".format(xls_name))

            wb.save(xls_path)

            return SUCCESS_STATUS, xls_path

        except Exception as e:
            return FAILURE_STATUS, repr(e)


class DataExportTxt(object):
    pass
