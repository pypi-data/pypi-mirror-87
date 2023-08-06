#
#   Copyright (c) 2014-2016, 2018-2020 eGauge Systems LLC
# 	1644 Conestoga St, Suite 2
# 	Boulder, CO 80301
# 	voice: 720-545-9767
# 	email: davidm@egauge.net
#
#   All rights reserved.
#
#   This code is the property of eGauge Systems LLC and may not be
#   copied, modified, or disclosed without any prior and written
#   permission from eGauge Systems LLC.
#
# pylint: disable=no-member
from datetime import date, datetime

from django.conf import settings as cfg
from django.contrib.auth.decorators import permission_required
from django.contrib.humanize.templatetags.humanize import intcomma
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404

import django.utils.text

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, fills
from openpyxl.utils import get_column_letter

from epic import perms
from epic.lib import get_stock_summary, get_inventory_summary
from epic.models import Inventory, Line_Item, Order, Part, Shipment, \
    Vendor, Vendor_Part, Warehouse

FONT_NAME = 'Arial'
FONT_SIZE = 10
FMT_COUNT = '#,##0'		# integer
FMT_PRICE = '$ #,##0.000000'	# currency w/ 6 decimals
FMT_AMOUNT = '$ #,##0.00'	# currency w/ 2 decimals
# integer show negative numbers with red:
FMT_DIFF = '[Green]#,##0;[Red]-#,##0;0'

def slugify(string):
    '''Like Django's slugify, but turns colons into dashes so that
    "bom:hv-sensor" turns into "bom-hv-sensor" rather than
    "bomhv-sensor", for example.

    '''
    string = string.replace(':', '-')
    return django.utils.text.slugify(string, allow_unicode=True)

class Style():
    fields = ('font', 'fill', 'border', 'alignment', 'number_format',
              'protection')

    def __init__(self, **kwargs):
        for f in Style.fields:
            if f in kwargs:
                self.__dict__[f] = kwargs[f]

    def copy(self, **kwargs):
        ns = Style()
        for f in Style.fields:
            if f in kwargs:
                ns.__dict__[f] = kwargs[f]
            elif f in self.__dict__:
                ns.__dict__[f] = self.__dict__[f]
        return ns

    def apply(self, target):
        if 'font' in self.__dict__:
            target.font = self.font
        if 'fill' in self.__dict__:
            target.fill = self.fill
        if 'border' in self.__dict__:
            target.border = self.border
        if 'alignment' in self.__dict__:
            target.alignment = self.alignment
        if 'number_format' in self.__dict__:
            target.number_format = self.number_format
        if 'protection' in self.__dict__:
            target.protection = self.protection

class WorksheetWriter():

    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.row = 1
        self.col = 0

    def set_row(self, row):
        self.row = row
        self.col = 0

    def next_row(self):
        self.set_row(self.row + 1)

    def append(self, val, style=None, col=None):
        if col is None:
            self.col += 1
            col = self.col
        else:
            self.col = col
        c = self.worksheet.cell(row=self.row, column=col)
        if val is not None:
            c.value = val
        if style is not None:
            style.apply(c)

def render_xls(filename, workbook):
    """Render the openpyxl WORKBOOK spreadsheet into an HTTP response.
    FILENAME is the name that will be given to the downloaded
    spreadsheet (sans .xlsx extension).
    """
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % filename
    workbook.save(response)
    return response

def append_vendor_parts(vendors, ww, part):
    vps = Vendor_Part.objects.filter(part_id=part.id)
    for v in vendors:
        vpn = ''
        for vp in vps:
            if vp.vendor.id == v.id:
                vpn = vp.vendor_pn
                break
        ww.append(vpn)

@permission_required(perms.VIEW)
def part_bom_export(request, pk):
    # pylint: disable=unused-argument
    part = get_object_or_404(Part, pk=pk)

    assy_items = part.assembly_items()
    if not assy_items:
        raise Http404	# not an assembly

    wb = Workbook()
    ws = wb.active
    # worksheet title must not be longer than 31 characters:
    ws.title = slugify(part.mfg_pn)[:31]
    ww = WorksheetWriter(ws)

    gray90 = PatternFill(fill_type=fills.FILL_SOLID, start_color='ffe5e5e5')
    gray95 = PatternFill(fill_type=fills.FILL_SOLID, start_color='fff2f2f2')

    style_normal = Style(font=Font(FONT_NAME, FONT_SIZE))
    style_comment = Style(font=Font(FONT_NAME, FONT_SIZE, italic=True))
    style_title = Style(font=Font(FONT_NAME, FONT_SIZE, bold=True))
    style_substitute = [
        Style(font=Font(FONT_NAME, FONT_SIZE), fill=gray90),
        Style(font=Font(FONT_NAME, FONT_SIZE), fill=gray95),
    ]

    pids = [item.comp.id for item in assy_items]
    vendors = Vendor.objects.filter(vendor_part__part_id__in=pids).distinct()

    cols = [('Idx', 6), ('PN', 10), ('Substitute PN', 13),
            ('Manufacturer', 24), ('Manufacturer PN', 31),
            ('Value', 8), ('Qty', 8),
            ('Price', 12), ('Refdes', 64), ('Description', 24)]
    for v in vendors:
        cols.append((v.name, 24))
    for cspec in cols:
        ww.append(cspec[0])
        ws.column_dimensions[get_column_letter(ww.col)].width = cspec[1]
    style_title.apply(ws.row_dimensions[ww.row])
    ww.next_row()

    sub_style_idx = 0
    idx = 1
    for item in assy_items:
        best_part = item.comp.best_part()
        refdes = ', '.join(sorted(item.refdes.split(',')))
        equiv_parts = best_part.equivalent_parts()

        if len(equiv_parts) > 1:
            style = style_substitute[sub_style_idx]
            sub_style_idx = (sub_style_idx + 1) % 2
        else:
            style = style_normal

        ww.append(idx)
        ww.append(str(best_part))
        ww.append(None)		# skip over Substitute PN
        ww.append(best_part.mfg)
        ww.append(best_part.mfg_pn)
        ww.append(best_part.val)
        ww.append(item.qty)
        ww.append(best_part.target_price,
                  style.copy(number_format='$ #,##0.000'))
        ww.append(refdes)
        ww.append(best_part.descr)
        style.apply(ws.row_dimensions[ww.row])
        append_vendor_parts(vendors, ww, best_part)
        ww.next_row()

        # add substitute rows, if any:
        for sub in equiv_parts:
            if sub == best_part or sub.status == Part.STATUS_PREVIEW:
                continue
            ww.append(str(sub), style, col=3)
            ww.append(sub.mfg, style)
            ww.append(sub.mfg_pn, style)
            ww.append(sub.val, style)
            ww.append(None, style)	# skip over Qty
            ww.append(None, style)	# skip over Price
            ww.append(None, style)	# skip over Refdes
            ww.append(sub.descr, style)
            append_vendor_parts(vendors, ww, sub)
            ww.next_row()
        idx += 1
    ww.next_row()
    ww.append('End of BOM', style_comment)
    return render_xls(slugify(ws.title), wb)

@permission_required(perms.VIEW)
def order_export(request, pk):
    # pylint: disable=unused-argument
    order = get_object_or_404(Order, pk=pk)
    line_items = Line_Item.objects.filter(txtn_id=pk)

    wb = Workbook()
    ws = wb.active
    ws.title = 'PO# %d' % order.id
    ww = WorksheetWriter(ws)

    style_normal = Style(font=Font(FONT_NAME, FONT_SIZE))
    style_comment = Style(font=Font(FONT_NAME, FONT_SIZE, italic=True))
    style_title = Style(font=Font(FONT_NAME, FONT_SIZE, bold=True))
    style_qty = style_normal.copy(number_format=FMT_COUNT)
    style_part_cost = style_normal.copy(number_format=FMT_PRICE)
    style_line_cost = style_normal.copy(number_format=FMT_AMOUNT)

    ww.append('Purchase Order #', style_title)
    ww.append('%d' % order.id, style_normal)
    ww.next_row()
    ww.append('Vendor', style_title)
    ww.append(order.vendor.name, style_normal)
    ww.next_row()
    ww.append('Order Date', style_title)
    ww.append(order.ts.strftime('%b %d, %Y'), style_normal)
    ww.next_row()
    ww.append('Expected Arrival Date', style_title)
    ww.append(order.expected_arrival_date.strftime('%b %d, %Y'), style_normal)
    ww.next_row()
    ww.next_row()

    ww.append('Bill To', style_title)
    for l in cfg.EPIC_BILL_TO_ADDRESS.split('\n'):
        ww.append(l, style_normal, col=2)
        ww.next_row()
    ww.next_row()

    ww.append('Ship To', style_title)
    for l in order.warehouse.address.split('\n'):
        ww.append(l, style_normal, col=2)
        ww.next_row()
    ww.next_row()

    ww.append('Ship By', style_title)
    ww.append(cfg.EPIC_SHIPPING_TYPE, style_normal)
    ww.next_row()

    ww.append('Shipping Account', style_title)
    ww.append(cfg.EPIC_SHIPPING_ACCOUNT, style_normal)
    ww.next_row()
    ww.next_row()

    cols = [('Idx', 20), ('Qty', 20), ('PN', 10), ('Vendor PN', 31),
            ('Manufacturer', 24), ('Manufacturer PN', 31),
            ('Price', 16), ('Amount', 16)]
    for cspec in cols:
        ww.append(cspec[0])
        ws.column_dimensions[get_column_letter(ww.col)].width = cspec[1]
    style_title.apply(ws.row_dimensions[ww.row])
    ww.next_row()

    total = 0
    for item in line_items:
        vendor_part = Vendor_Part.get(item.part.id, order.vendor.id)
        total += item.line_cost
        ww.append(item.index)
        ww.append(item.qty, style_qty)
        ww.append(str(item.part))
        ww.append(vendor_part.vendor_pn)
        ww.append(item.part.mfg)
        ww.append(item.part.mfg_pn)
        ww.append(item.part_cost(), style_part_cost)
        ww.append(item.line_cost, style_line_cost)
        style_normal.apply(ws.row_dimensions[ww.row])
        ww.next_row()
    ww.append('Order Total', style_title, col=7)
    ww.append(total, style_title.copy(number_format=FMT_AMOUNT))
    ww.next_row()
    ww.append('End of Order', style_comment)
    return render_xls('order-%d-%s' % (order.id, slugify(order.vendor.name)),
                      wb)

@permission_required(perms.VIEW)
def ship_export(request, pk):
    # pylint: disable=unused-argument
    ship = get_object_or_404(Shipment, pk=pk)
    line_items = Line_Item.objects.filter(txtn_id=pk)

    wb = Workbook()
    ws = wb.active
    ws.title = 'SO# %d' % ship.id
    ww = WorksheetWriter(ws)

    style_normal = Style(font=Font(FONT_NAME, FONT_SIZE))
    style_comment = Style(font=Font(FONT_NAME, FONT_SIZE, italic=True))
    style_title = Style(font=Font(FONT_NAME, FONT_SIZE, bold=True))
    style_qty = style_normal.copy(number_format=FMT_COUNT)

    ww.append('Ship Order #', style_title)
    ww.append('%d' % ship.id, style_normal)
    ww.next_row()

    ww.append('Tracking #', style_title)
    ww.append(ship.tracking, style_normal)
    ww.next_row()

    ww.append('Ship Date', style_title)
    ww.append(ship.ts.strftime('%b %d, %Y'), style_normal)
    ww.next_row()
    ww.append('Ship From', style_title)
    ww.append(ship.ordr.vendor.name if ship.ordr
              else ship.from_warehouse.name, style_normal)
    ww.next_row()
    ww.next_row()

    ww.append('Ship To', style_title)
    for l in ship.warehouse.address.split('\n'):
        ww.append(l, style_normal, col=2)
        ww.next_row()
    ww.next_row()

    # Don't include any price information in the ship order.  That info
    # is somewhat sensitive and kept best between the vendor and ourselves.
    # We don't want to share that with the assembly houses.
    cols = [('Idx', 20), ('Qty', 20), ('PN', 10), ('Vendor PN', 31),
            ('Manufacturer', 24), ('Manufacturer PN', 31)]
    for cspec in cols:
        if ship.ordr is None and cspec[0] == 'Vendor PN':
            continue	# no vendor for inter-warehouse shipments
        ww.append(cspec[0])
        ws.column_dimensions[get_column_letter(ww.col)].width = cspec[1]
    style_title.apply(ws.row_dimensions[ww.row])
    ww.next_row()

    for item in line_items:
        ww.append(item.index)
        ww.append(item.qty, style_qty)
        ww.append(str(item.part))
        if ship.ordr:
            vendor_part = Vendor_Part.get(item.part.id, ship.ordr.vendor.id)
            ww.append(vendor_part.vendor_pn)
        ww.append(item.part.mfg)
        ww.append(item.part.mfg_pn)
        style_normal.apply(ws.row_dimensions[ww.row])
        ww.next_row()
    ww.next_row()
    ww.append('End of Shipment', style_comment)
    return render_xls('ship-%d-%s' % (ship.id, slugify(ship.warehouse.name)),
                      wb)

@permission_required(perms.VIEW)
def warehouse_stock_export(request, pk):

    def part_ref(part_list, index):
        result = str(part_list[index][0])
        if len(part_list) > 1:
            result += ' * %s' % intcomma(part_list[index][1])
        return result

    def row_style(style, index):
        if index % 2 == 0:
            return style.copy(fill=gray95)
        return style.copy(fill=gray90)

    if 'as_of_date' in request.GET:
        as_of_date = datetime.strptime(request.GET['as_of_date'], '%Y-%m-%d') \
                             .date()
    else:
        as_of_date = date.today()

    if pk is None:
        warehouse = None
    else:
        warehouse = get_object_or_404(Warehouse, pk=pk)
    stock, stock_total = get_stock_summary(warehouse, as_of_date=as_of_date)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Summary '
    ww = WorksheetWriter(ws)

    gray90 = PatternFill(fill_type=fills.FILL_SOLID, start_color='ffe5e5e5')
    gray95 = PatternFill(fill_type=fills.FILL_SOLID, start_color='fff2f2f2')
    style_normal = Style(font=Font(FONT_NAME, FONT_SIZE))
    style_comment = Style(font=Font(FONT_NAME, FONT_SIZE, italic=True))
    style_title = Style(font=Font(FONT_NAME, FONT_SIZE, bold=True))
    style_qty = style_normal.copy(number_format=FMT_COUNT)

    title = 'Stock as of %s ' % as_of_date.strftime('%b %d, %Y')
    if pk is None:
        title += 'for all warehouses together'
    else:
        title += 'at warehouse %s' % warehouse.name

    ww.append(title, style_title)
    ww.next_row()
    ww.next_row()

    cols = [('Idx', 6), ('Qty', 10), ('PN', 15),
            ('Manufacturer', 24), ('Manufacturer PN', 31),
            ('Price', 16), ('Amount', 16)]
    for cspec in cols:
        ww.append(cspec[0])
        ws.column_dimensions[get_column_letter(ww.col)].width = cspec[1]
    style_title.apply(ws.row_dimensions[ww.row])
    ww.next_row()

    index = 1
    for row in stock:
        style = row_style(style_normal, index)
        ww.append(index, style)
        ww.append(row['total_qty'], row_style(style_qty, index))
        part_list = row['part_list']
        part = part_list[0][0]
        ww.append(part_ref(part_list, 0), style)
        ww.append(part.mfg, style)
        ww.append(part.mfg_pn, style)
        ww.append(row['price'], style.copy(number_format=FMT_PRICE))
        ww.append(row['amount'], style.copy(number_format=FMT_AMOUNT))
        for i in range(1, len(part_list)):
            ww.next_row()
            part = part_list[i][0]
            ww.append(None, style)
            ww.append(None, style)
            ww.append(part_ref(part_list, i), style)
            ww.append(part.mfg, style)
            ww.append(part.mfg_pn, style)
            ww.append(None, style)
            ww.append(None, style)
        ww.next_row()
        index += 1
    ww.append('Parts Total', style_title, col=6)
    ww.append(stock_total, style_title.copy(number_format=FMT_AMOUNT))
    ww.next_row()
    ww.append('End of Report', style_comment)

    return render_xls('stock-%s' % (slugify(warehouse.name) if warehouse
                                    else 'all'),
                      wb)

@permission_required(perms.VIEW)
def warehouse_stock_all_export(request):
    return warehouse_stock_export(request, None)

@permission_required(perms.VIEW)
def warehouse_inv_export(request, warehouse, pk):
    # pylint: disable=unused-argument

    def row_style(style, index):
        if index % 2 == 0:
            return style.copy(fill=gray95)
        return style.copy(fill=gray90)

    inventory = get_object_or_404(Inventory, pk=pk)
    warehouse = get_object_or_404(Warehouse, pk=warehouse)
    if inventory.warehouse_id != warehouse.id:
        raise Http404

    ret = get_inventory_summary(warehouse.id, inventory)
    inv_items = ret[0]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory # %d ' % inventory.id
    ww = WorksheetWriter(ws)

    gray90 = PatternFill(fill_type=fills.FILL_SOLID, start_color='ffe5e5e5')
    gray95 = PatternFill(fill_type=fills.FILL_SOLID, start_color='fff2f2f2')
    style_normal = Style(font=Font(FONT_NAME, FONT_SIZE))
    style_comment = Style(font=Font(FONT_NAME, FONT_SIZE, italic=True))
    style_title = Style(font=Font(FONT_NAME, FONT_SIZE, bold=True))
    style_qty = style_normal.copy(number_format=FMT_COUNT)

    title = 'Inventory as of %s for warehouse %s' \
            % (inventory.ts.strftime('%b %d, %Y'), warehouse.name)

    ww.append(title, style_title)
    ww.next_row()
    ww.next_row()

    # Don't include value change so this spreadsheet can be shared with
    # assembly-houses:
    cols = [('Idx', 6), ('PN', 15),
            ('Manufacturer', 24), ('Manufacturer PN', 31),
            ('New Qty', 10), ('Old Qty', 10), ('Qty Change', 10),
            ('Net Change', 10)]
    for cspec in cols:
        ww.append(cspec[0])
        ws.column_dimensions[get_column_letter(ww.col)].width = cspec[1]
    style_title.apply(ws.row_dimensions[ww.row])
    ww.next_row()

    index = 1
    for row in inv_items:
        style = row_style(style_normal, index)
        ww.append(index, style)
        qstyle = row_style(style_qty, index)
        net_dif = 0
        for i in range(len(row['qty_dif_list'])):
            net_dif += row['qty_dif_list'][i][0]
        for i in range(len(row['qty_new_list'])):
            if i > 0:
                ww.next_row()
                ww.append('', style)	# write empty index column
            part = row['qty_new_list'][i][1]
            ww.append(str(part), style)
            ww.append(part.mfg)
            ww.append(part.mfg_pn)
            ww.append(row['qty_new_list'][i][0], qstyle)
            ww.append(row['qty_old_list'][i][0], qstyle)
            ww.append(row['qty_dif_list'][i][0], qstyle)
            if i == 0:
                ww.append(net_dif, qstyle)
            else:
                ww.append('', qstyle)
        ww.next_row()
        index += 1
    ww.next_row()
    ww.append('End of Inventory', style_comment)

    return render_xls('inventory-%d-%s' % (inventory.id,
                                           slugify(warehouse.name)),
                      wb)
