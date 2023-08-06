"""
microsoft.py
written in Python3
author: C. Lockhart <chris@lockhartlab.org>
"""

import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.drawing.image import Image
from openpyxl.styles import Font


# Class to write to Excel
class Excel:
    """
    Write to Excel
    """

    # Initialize class instance
    def __init__(self, filename, option='w'):
        """
        Initialize class instance

        Parameters
        ----------
        filename : str
            Name of workbook
        """

        # Name of file
        self.filename = filename

        # Option
        if option not in ['r', 'w']:
            raise AttributeError('option not understood')

        # If option = 'r', read in file
        if option == 'r':
            self.workbook = load_workbook(filename)


# Helper function to embolden cell
def _bold(worksheet, rid, cid):
    """
    Change formatting of cell to `bold`

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet
    rid : int
        Row index
    cid : int
        Column index
    """

    worksheet.cell(row=rid, column=cid).font = Font(bold=True)


# Helper function to decode A1 notation into row and column indices
def _decode_a1(a1):
    """
    Decode A1 notation to row and column indices

    Parameters
    ----------
    a1 : str
        Cell location in A1 notation

    Returns
    -------
    int, int
        Row index and column index
    """

    # Get coordinate from A1 notation
    coord = coordinate_from_string(a1)

    # Separate into row ID, column ID
    rid = coord[1]
    cid = column_index_from_string(coord[0])

    # Return
    return rid, cid


# Write DataFrame to cell
def _write_frame(worksheet, a1, df, header=True, bold_header=False, bold_index=False):
    # Get starting rid, cid from A1
    rid_start, cid_start = _decode_a1(a1)

    # Copy DataFrame and turn into an array
    df = df.copy().reset_index()
    rows = df.values
    if header:
        rows = np.vstack([df.columns, rows])

    # Iterate over rows and columns, writing to cells
    for rid, row in enumerate(rows, rid_start):
        for cid, value in enumerate(row, cid_start):
            if isinstance(value, np.ndarray):
                value = np.reshape(value, -1)[0]
            worksheet.cell(row=rid, column=cid, value=value)
            if (bold_header & (rid == rid_start)) | (bold_index & (cid == cid_start)):
                _bold(worksheet, rid, cid)


# Write image to cell
def _write_image(worksheet, a1, image_path, width_scale=0.75, height_scale=0.75):
    """
    Write image to worksheet cell

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet object to write to
    a1 : str
        Cell in A1 notation
    image_path : str
        Local path to image
    width_scale : float
        Value to scale the image width (Default: 0.75)
    height_scale : float
        Value to scale the image height (Default: 0.75)
    """

    # Create image
    image = Image(image_path)

    # Scale
    image.width *= width_scale
    image.height *= height_scale

    # Anchor the image to A1
    image.anchor(worksheet[a1])

    # Add image to the worksheet
    worksheet.add_image(image)


# Write value to cell
def _write_value(worksheet, a1, value, bold=False):
    """
    Write value to worksheet cell

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet object to write to
    a1 : str
        Cell in A1 notation
    value : any
        Value to write
    bold : bool
        Should the value be bolded? (Default: False)
    """

    # Separate out column and row from cell (in A1 notation)
    rid, cid = _decode_a1(a1)

    # Write to cell
    worksheet.cell(row=rid, column=cid, value=value)

    # Format
    if bold:
        _bold(worksheet, rid, cid)
